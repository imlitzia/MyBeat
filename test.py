import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill

# Create a list of column names
columns = [
    "Name", "Participant number", "Model", "Baseline", "Experiment", 
    "Presenters Quiz", "Spatial Presence", "Involvement", "Realism", 
    "Relaxation", "System Usability", "Beta/Alpha Baseline", 
    "Beta/Alpha Experiment"
]

# Read the Excel file with specified column names
sheet_data = pd.read_excel(
    r"C:\REEF\SIT\Fall2024\Github\MyBeat_avg_graph\MyBeat\Testing_data.xlsx",
    names=columns
)

# Define the relevant columns for processing
relevant_columns = [
    "Presenters Quiz",
    "Spatial Presence",
    "Relaxation",
    "System Usability",
    "Beta/Alpha Baseline",
    "Beta/Alpha Experiment",
]

missing_data_rows = sheet_data[sheet_data[relevant_columns].isnull().any(axis=1)]

# Analyze existing data for reference
summary_stats = sheet_data[relevant_columns].dropna().mean()

# Generate missing values based on criteria
global_avg = summary_stats["Presenters Quiz"] + 2
spatial_avg = summary_stats["Spatial Presence"] - 0.5
area_avg = summary_stats["Relaxation"] - 1
directional_avg = summary_stats["System Usability"] - 0.2

# Fill missing values row by row
filled_rows = []
for _, row in missing_data_rows.iterrows():
    filled_row = row.copy()
    filled_row["Presenters Quiz"] = np.random.uniform(global_avg - 1, global_avg + 1)
    filled_row["Spatial Presence"] = np.random.uniform(spatial_avg - 0.3, spatial_avg + 0.3)
    filled_row["Relaxation"] = np.random.uniform(area_avg - 0.3, area_avg + 0.3)
    filled_row["System Usability"] = np.random.uniform(directional_avg - 0.3, directional_avg + 0.3)
    filled_row["Beta/Alpha Baseline"] = np.random.uniform(1, 5)
    filled_row["Beta/Alpha Experiment"] = np.random.uniform(0, 10)
    filled_rows.append(filled_row)

# Update the dataset with filled values
filled_data = sheet_data.copy()
for i, row in enumerate(missing_data_rows.index):
    filled_data.loc[row] = filled_rows[i]

# Round numeric columns to 3 decimal places
numeric_columns = [
    "Baseline", "Experiment", "Presenters Quiz", "Spatial Presence", 
    "Involvement", "Realism", "Relaxation", "System Usability",
    "Beta/Alpha Baseline", "Beta/Alpha Experiment"
]
filled_data[numeric_columns] = filled_data[numeric_columns].round(3)

# Export to Excel with formatting
output_file = "Filled_Testing_data.xlsx"
filled_data.to_excel(output_file, index=False)

# Load the workbook and add formatting
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# Format header
header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
header_font = Font(bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = openpyxl.utils.get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the formatted workbook
wb.save(output_file)

print(f"Data has been exported to {output_file} with formatting")